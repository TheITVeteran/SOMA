import sys
import json
import openpyxl
import os
import pandas as pd

def analyze_structure(file_path):
    """
    Parses Excel structure with aggregated risk scoring.
    """
    wb = openpyxl.load_workbook(file_path, data_only=False)
    findings = []
    sheet_stats = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        overrides = 0
        errors = 0
        total_cells = 0
        
        for row in ws.iter_rows():
            for cell in row:
                total_cells += 1
                
                # 1. Detect Hardcoded Overrides in Formula-Heavy Columns
                if cell.value is not None and not isinstance(cell.value, (str, openpyxl.formula.Formula)):
                    # Production-grade check: Neighbors
                    col_idx = cell.column
                    formula_neighbors = 0
                    test_range = 10
                    for i in range(1, test_range + 1):
                        test_cell = ws.cell(row=max(1, cell.row + (i - test_range//2)), column=col_idx)
                        if test_cell.data_type == 'f':
                            formula_neighbors += 1
                    
                    if formula_neighbors > 4: # 50%+ are formulas
                        overrides += 1
                        findings.append({
                            "sheet": sheet_name,
                            "cell": cell.coordinate,
                            "type": "HARDCODED_OVERRIDE",
                            "value": cell.value,
                            "severity": "CRITICAL",
                            "context": f"Found in a column where {formula_neighbors}/{test_range} nearby cells are formulas."
                        })

                # 2. Detect Excel Errors
                if isinstance(cell.value, str) and cell.value.startswith('#'):
                    errors += 1
                    findings.append({
                        "sheet": sheet_name,
                        "cell": cell.coordinate,
                        "type": "FORMULA_ERROR",
                        "value": cell.value,
                        "severity": "HIGH"
                    })

        sheet_stats[sheet_name] = {
            "overrides": overrides,
            "errors": errors,
            "risk_density": (overrides + errors) / max(1, total_cells)
        }

    return {
        "success": True,
        "total_findings": len(findings),
        "findings": findings,
        "sheet_stats": sheet_stats,
        "overall_risk_score": sum(s["risk_density"] for s in sheet_stats.values()) / len(sheet_stats)
    }

def main():
    try:
        raw_input = sys.stdin.read().strip()
        if not raw_input: return
        cmd = json.loads(raw_input)
        file_path = cmd.get("input")

        if not os.path.exists(file_path):
            print(json.dumps({"success": False, "error": "File not found"}))
            return

        result = analyze_structure(file_path)
        print(json.dumps(result))

    except Exception as e:
        print(json.dumps({"success": False, "error": str(e)}))

if __name__ == "__main__":
    main()
